import openpyxl, json

wb = openpyxl.load_workbook('BLOG/GetCourse/SNP/KontentMatritsa/Контент-матрица для блога.xlsx')

data = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows_data = []
    current_category = None
    current_trigger = None

    has_niche = len(headers) >= 5 and headers[2] == 'Ниша'

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        max_check = 5 if has_niche else 4
        if not any(cell is not None for cell in row_list[:max_check]):
            continue

        if row_list[0] is not None:
            has_other = any(cell is not None for cell in row_list[1:max_check])
            if not has_other:
                current_category = row_list[0]
                continue
            else:
                current_trigger = row_list[0]

        if has_niche:
            entry = {
                'category': current_category,
                'trigger': current_trigger if current_trigger else '',
                'topic': row_list[1] or '',
                'niche': row_list[2] or '',
                'idea': row_list[3] or '',
            }
        else:
            entry = {
                'category': current_category,
                'trigger': current_trigger if current_trigger else '',
                'topic': row_list[1] or '',
                'idea': row_list[2] or '',
            }

        if entry.get('idea') or entry.get('topic'):
            rows_data.append(entry)

    data[sheet_name] = rows_data
    print(f'{sheet_name}: {len(rows_data)} entries')

with open('BLOG/GetCourse/SNP/KontentMatritsa/matrix_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print('Saved to matrix_data.json')
